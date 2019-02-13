# Import excel module and csv
from openpyxl import load_workbook
import csv
import sys

# Open XL
wb = load_workbook(filename="iris_data.xlsx")
ws = wb["Fisher's Iris Data"]


# Define headers
def read_headers():
    headers = []
    # Read first row(for headers)
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            headers.append(cell.value)

    return headers


def read_data():
    # Create array of dictionary to store values
    data = []

    # Read the rest of the row, and put them in a array of dictionarys
    for row in ws.iter_rows(min_row=2):
        data.append(
            {"Sepal length": row[0].value, "Sepal width": row[1].value, "Petal length": row[2].value, "Petal width": row[3].value, "Species": row[4].value})
    return data


def write_to_csv():
    # Get info
    csvHeaders = read_headers()
    csvData = read_data()

    # Create csv file
    with open('iris_data.csv', mode='w') as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=csvHeaders)

        # Write headers to csv file
        writer.writeheader()

        # Loop through each row, and insert into the csv file
        for i in range(len(csvData)):
            writer.writerow(csvData[i])

        print("CSV file created! Go check ./iris_data.csv")

        



def test():
    # Test headers
    expectedHeaders = ["Sepal length", "Sepal width",
                       "Petal length", "Petal width", "Species"]
    headers = read_headers()

    if(headers == expectedHeaders):
        print("Test headers success")
    else:
        print("expected:" + expectedHeaders + "  But got: " + headers)

    # Test data
    expected_first_line = {"Sepal length": 5.10, "Sepal width": 3.50,
                           "Petal length": 1.40, "Petal width": 0.20, "Species": "I. setosa"}

    actual_first_line = read_data()

    if(expected_first_line == actual_first_line[0]):
        print("Test data Sucess")
    else:
        print("expected:" + expected_first_line +
              "  But got: " + actual_first_line[0])


if(sys.argv[1] == "test"):
    test()
elif(sys.argv[1] == "run"):
    write_to_csv()
else:
    print("Please input a argument when running the program! (either 'run' or 'test')")
